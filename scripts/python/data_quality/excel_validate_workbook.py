#!/usr/bin/env python3
"""Validate Excel workbook sheets, required columns, and allowed values."""

from __future__ import annotations

import argparse
import json
from pathlib import Path


def parse_sheet_columns(entries: list[str]) -> dict[str, list[str]]:
    parsed: dict[str, list[str]] = {}
    for entry in entries:
        # Format: SheetName:col1,col2,col3
        if ":" not in entry:
            continue
        sheet, cols = entry.split(":", 1)
        parsed[sheet.strip()] = [c.strip() for c in cols.split(",") if c.strip()]
    return parsed


def parse_allowed_values(entries: list[str]) -> dict[tuple[str, str], set[str]]:
    parsed: dict[tuple[str, str], set[str]] = {}
    for entry in entries:
        # Format: SheetName:ColumnName:val1|val2|val3
        parts = entry.split(":", 2)
        if len(parts) != 3:
            continue
        sheet, column, values = parts
        parsed[(sheet.strip(), column.strip())] = {v.strip() for v in values.split("|") if v.strip()}
    return parsed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate Excel workbook structure and data rules.")
    parser.add_argument("--workbook", required=True, help="Path to .xlsx workbook")
    parser.add_argument(
        "--required-sheets",
        nargs="*",
        default=[],
        help="List of required sheet names",
    )
    parser.add_argument(
        "--required-columns",
        nargs="*",
        default=[],
        help="Rules like Sheet:col1,col2",
    )
    parser.add_argument(
        "--allowed-values",
        nargs="*",
        default=[],
        help="Rules like Sheet:Column:open|closed",
    )
    parser.add_argument("--output", required=True, help="Path to write validation JSON")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook)
    output_path = Path(args.output)

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for workbook validation") from exc

    wb = load_workbook(workbook_path, data_only=True)

    required_cols_map = parse_sheet_columns(args.required_columns)
    allowed_values_map = parse_allowed_values(args.allowed_values)

    issues: list[dict[str, str]] = []

    for sheet in args.required_sheets:
        if sheet not in wb.sheetnames:
            issues.append({"type": "missing_sheet", "sheet": sheet, "detail": "Sheet not found"})

    for sheet_name, expected_cols in required_cols_map.items():
        if sheet_name not in wb.sheetnames:
            issues.append(
                {
                    "type": "missing_sheet",
                    "sheet": sheet_name,
                    "detail": "Required for column validation",
                }
            )
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        headers_norm = [str(h).strip() if h is not None else "" for h in headers]

        for col in expected_cols:
            if col not in headers_norm:
                issues.append(
                    {
                        "type": "missing_column",
                        "sheet": sheet_name,
                        "detail": col,
                    }
                )

    for (sheet_name, column_name), allowed in allowed_values_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        if column_name not in headers:
            continue

        col_index = headers.index(column_name) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_index).value
            text_val = "" if cell_val is None else str(cell_val).strip()
            if text_val and text_val not in allowed:
                issues.append(
                    {
                        "type": "invalid_value",
                        "sheet": sheet_name,
                        "detail": f"{column_name} row {row_idx}: {text_val}",
                    }
                )

    result = {
        "workbook": str(workbook_path),
        "issue_count": len(issues),
        "issues": issues,
        "valid": len(issues) == 0,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, indent=2), encoding="utf-8")

    print(f"Workbook validation complete: issues={len(issues)} -> {output_path}")
    return 0 if result["valid"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
