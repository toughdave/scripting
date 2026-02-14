#!/usr/bin/env python3
"""Build admissions/results audit packets from CSV extracts for review and sign-off cycles."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any


def parse_float(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export audit packet artifacts (clean dataset, anomaly register, summary, optional workbook) "
            "for operational review workflows."
        )
    )
    parser.add_argument("--input", required=True, help="Input CSV path")
    parser.add_argument("--output-dir", required=True, help="Audit packet directory")
    parser.add_argument("--key-column", default="student_id", help="Primary identity column")
    parser.add_argument("--score-column", default="score", help="Numeric score column")
    parser.add_argument("--low-score-threshold", type=float, default=50.0, help="Low score threshold")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    out_dir = Path(args.output_dir)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    with input_path.open("r", newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    fieldnames = list(rows[0].keys()) if rows else []
    anomalies: list[dict[str, Any]] = []

    for row in rows:
        key_val = (row.get(args.key_column) or "").strip()
        score_val = (row.get(args.score_column) or "").strip()
        score_num = parse_float(score_val)

        if not key_val:
            issue = dict(row)
            issue["issue"] = f"missing_{args.key_column}"
            anomalies.append(issue)

        if score_val == "":
            issue = dict(row)
            issue["issue"] = f"missing_{args.score_column}"
            anomalies.append(issue)
        elif score_num is None:
            issue = dict(row)
            issue["issue"] = f"invalid_{args.score_column}"
            anomalies.append(issue)
        elif score_num < args.low_score_threshold:
            issue = dict(row)
            issue["issue"] = "low_score"
            anomalies.append(issue)

    out_dir.mkdir(parents=True, exist_ok=True)
    cleaned_csv = out_dir / "cleaned.csv"
    anomalies_csv = out_dir / "anomalies.csv"
    summary_json = out_dir / "summary.json"
    workbook_path = out_dir / "audit_packet.xlsx"

    write_csv(cleaned_csv, rows, fieldnames)
    anomaly_fields = fieldnames + ["issue"] if fieldnames else ["issue"]
    write_csv(anomalies_csv, anomalies, anomaly_fields)

    summary = {
        "input": str(input_path),
        "rows": len(rows),
        "anomaly_rows": len(anomalies),
        "key_column": args.key_column,
        "score_column": args.score_column,
        "low_score_threshold": args.low_score_threshold,
        "artifacts": {
            "cleaned_csv": str(cleaned_csv),
            "anomalies_csv": str(anomalies_csv),
            "summary_json": str(summary_json),
            "workbook": str(workbook_path),
        },
    }
    summary_json.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    workbook_written = False
    try:
        from openpyxl import Workbook  # type: ignore

        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Data"
        if fieldnames:
            ws_data.append(fieldnames)
            for row in rows:
                ws_data.append([row.get(col, "") for col in fieldnames])

        ws_anom = wb.create_sheet("Anomalies")
        ws_anom.append(anomaly_fields)
        for row in anomalies:
            ws_anom.append([row.get(col, "") for col in anomaly_fields])

        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["metric", "value"])
        ws_summary.append(["rows", len(rows)])
        ws_summary.append(["anomaly_rows", len(anomalies)])
        ws_summary.append(["low_score_threshold", args.low_score_threshold])

        wb.save(workbook_path)
        workbook_written = True
    except Exception:
        # Keep packet generation successful even when openpyxl is unavailable.
        workbook_written = False

    if workbook_written:
        print(f"Audit packet complete with workbook -> {out_dir}")
    else:
        print(f"Audit packet complete (CSV/JSON only) -> {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
